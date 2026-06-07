"""
ASX Energy NZ Futures Scraper
Scrapes Base Month and Base Quarter settle prices for Otahuhu and Benmore
and appends a row per contract per day to an Excel file.
"""

import sys
import time
import re
from datetime import datetime, date
from pathlib import Path
from zoneinfo import ZoneInfo

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

URL = "https://www.asxenergy.com.au/futures/nz_electricity"
EXCEL_FILE = Path(__file__).parent / "asx_nz_futures.xlsx"

TARGET_SECTIONS = {"Base Month", "Base Quarter"}

NODE_MAP = {
    "Otahuhu": "OTA2201",
    "Benmore": "BEN2201",
}

# Matches a section label div, e.g. "Otahuhu Base Month ED" / "Benmore Base Quarter EE"
LABEL_RE = re.compile(r"^(Otahuhu|Benmore)\s+(Base Month|Base Quarter|Peak Quarter|Base Cal)")

MONTHS = {m: i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], start=1)}


def get_driver():
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (X11; Linux x86_64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
    try:
        service = Service("/usr/bin/chromedriver")
        driver = webdriver.Chrome(service=service, options=opts)
    except Exception:
        try:
            from webdriver_manager.chrome import ChromeDriverManager
            driver = webdriver.Chrome(
                service=Service(ChromeDriverManager().install()), options=opts
            )
        except Exception:
            driver = webdriver.Chrome(options=opts)
    return driver


def clean_label(text):
    """Collapse whitespace and strip a trailing 1-3 letter junk code (e.g. 'Otahuhu Base Month ED')."""
    collapsed = re.sub(r"\s+", " ", text).strip()
    return re.sub(r"\s+[A-Z]{1,3}$", "", collapsed).strip()


def normalise_contract(raw):
    """
    Normalise the site's compact contract codes back to the long form used in
    the historical data, so the Excel file stays consistent over time.

      Month:   'Jun 26'  -> 'Jun 2026'
      Quarter: 'Q326'    -> 'Q3 2026'
    Anything unrecognised is returned unchanged.
    """
    raw = raw.strip()

    # Quarter: Q + quarter digit + 2-digit year, e.g. 'Q326'
    qm = re.match(r"^Q([1-4])\s*'?(\d{2})$", raw)
    if qm:
        q, yy = qm.group(1), qm.group(2)
        return f"Q{q} 20{yy}"

    # Month: 'Jun 26' or 'Jun26'
    mm = re.match(r"^([A-Za-z]{3})\s*'?(\d{2})$", raw)
    if mm and mm.group(1).title() in MONTHS:
        return f"{mm.group(1).title()} 20{mm.group(2)}"

    return raw


def scrape() -> list[dict]:
    """
    New page structure (Tailwind-based, confirmed live):
      <div class="... text-white ...">Otahuhu Base Month ED</div>
      <table> ... </table>     # contract is the FIRST <td>, Settle is the LAST <td>
      <div ...>Otahuhu Base Quarter EA</div>
      <table> ... </table>
      ...
    Data rows have a duplicated contract column (desktop + mobile responsive copies),
    so we always take the first <td> for the contract and the last <td> for Settle.
    """
    driver = get_driver()
    records = []

    try:
        driver.get(URL)

        try:
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "table"))
            )
        except TimeoutException:
            print("Timed out waiting for page to load", file=sys.stderr)
            return records

        time.sleep(5)

        from bs4 import BeautifulSoup
        soup = BeautifulSoup(driver.page_source, "html.parser")

        current_node = None
        current_section = None

        for elem in soup.find_all(["div", "table"]):

            if elem.name == "div":
                full = re.sub(r"\s+", " ", elem.get_text()).strip()
                # Only short label divs, not big container divs
                if len(full) >= 40:
                    continue
                m = LABEL_RE.match(clean_label(elem.get_text()))
                if m:
                    current_node = m.group(1)
                    current_section = m.group(2)
                continue

            # table
            if current_node is None or current_section not in TARGET_SECTIONS:
                continue

            rows = elem.find_all("tr")
            if len(rows) < 2:
                continue

            for row in rows[1:]:
                tds = row.find_all("td")
                if len(tds) < 3:
                    continue

                contract_raw = tds[0].get_text(strip=True)
                settle = tds[-1].get_text(strip=True)

                if not contract_raw or not re.search(r"\d", contract_raw):
                    continue

                settle_clean = settle.replace(",", "").strip()
                price = None
                if settle_clean not in ("-", "", "N/A", "n/a"):
                    try:
                        price = float(settle_clean)
                    except ValueError:
                        pass

                records.append({
                    "node": NODE_MAP.get(current_node, current_node),
                    "period_type": current_section,
                    "time_period": normalise_contract(contract_raw),
                    "price": price,
                })

            current_section = None  # consume so the next table needs its own label

    finally:
        driver.quit()

    return records


def append_to_excel(records: list[dict], execution_date: datetime):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    data_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_even = PatternFill("solid", start_color="EBF3FB")
    fill_odd = PatternFill("solid", start_color="FFFFFF")

    exec_date = execution_date.date()

    # Remove any existing rows for today to prevent duplicates
    rows_to_delete = []
    for row in ws.iter_rows(min_row=3):
        cell_value = row[0].value
        if cell_value is None:
            continue
        if isinstance(cell_value, datetime):
            row_date = cell_value.date()
        elif isinstance(cell_value, date):
            row_date = cell_value
        else:
            continue
        if row_date == exec_date:
            rows_to_delete.append(row[0].row)

    for row_num in reversed(rows_to_delete):
        ws.delete_rows(row_num)

    if rows_to_delete:
        print(f"Removed {len(rows_to_delete)} duplicate rows for {exec_date}")

    last_row = ws.max_row
    insert_row = 3 if last_row < 3 else last_row + 1

    for record in records:
        fill = fill_even if (insert_row % 2 == 0) else fill_odd
        cells_data = [
            exec_date,
            record["node"],
            record["period_type"],
            record["time_period"],
            record["price"],
        ]

        for col, value in enumerate(cells_data, 1):
            cell = ws.cell(row=insert_row, column=col, value=value)
            cell.font = data_font
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if col == 1:
                cell.number_format = "YYYY-MM-DD"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 5 and value is not None:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")

        insert_row += 1

    wb.save(EXCEL_FILE)
    print(f"Appended {len(records)} records for {exec_date}")


def main():
    nzt = ZoneInfo("Pacific/Auckland")
    execution_dt = datetime.now(tz=nzt)
    print(f"Scraping at {execution_dt.strftime('%Y-%m-%d %H:%M:%S %Z')}")

    records = scrape()

    if not records:
        print("No records scraped — check page structure or network access.", file=sys.stderr)
        sys.exit(1)

    print(f"Scraped {len(records)} records:")
    for r in records:
        print(f"  {r['node']:10} | {r['period_type']:14} | {r['time_period']:12} | {r['price']}")

    append_to_excel(records, execution_dt)


if __name__ == "__main__":
    main()
