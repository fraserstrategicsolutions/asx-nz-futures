import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
import time
import pytz

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup

DATASET_URL = "https://www.asxenergy.com.au/futures_nz/dataset"
EXCEL_FILE = "asx_nz_futures.xlsx"
LOCATIONS = ["Otahuhu", "Benmore"]
PRODUCT_TYPES = ["Base Month", "Base Quarter"]
CONTRACT_FIELDS = ["Bid Size", "Bid", "Ask", "Ask Size", "High", "Low", "Last", "+/-", "Vol"]


def fetch_html_with_selenium():
    """Use headless Chrome to fully render the JS-driven page."""
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    try:
        print(f"Loading {DATASET_URL} ...")
        driver.get(DATASET_URL)

        # Wait up to 30s for at least one table row to appear
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr"))
        )
        # Extra pause to let all sections finish rendering
        time.sleep(5)

        html = driver.page_source
        print(f"Page loaded — {len(html)} chars")
    finally:
        driver.quit()

    return html


def parse_data(html):
    """
    Parse rendered HTML and return:
      last_update (str)
      data: {location: {product_name: [ {Contract, Bid Size, ...}, ... ] }}
    """
    soup = BeautifulSoup(html, "html.parser")

    # Debug: show all text in #last div
    last_div = soup.find("div", id="last")
    last_update = ""
    if last_div:
        date_str = last_div.get("data-date", "")
        time_str = last_div.get("data-time", "")
        last_update = f"{date_str} {time_str}".strip()
        print(f"Found #last div: {last_update}")
    else:
        print("WARNING: #last div not found")

    # Debug: show all h2s found
    all_h2 = [h.get_text(strip=True) for h in soup.find_all("h2")]
    print(f"All h2 tags: {all_h2}")

    all_market = soup.find_all("div", class_="market-dataset")
    print(f"market-dataset divs found: {len(all_market)}")

    all_dataset = soup.find_all("div", class_="dataset")
    print(f"dataset divs found: {len(all_dataset)}")

    all_tables = soup.find_all("table")
    print(f"tables found: {len(all_tables)}")

    # If we find tables but no market-dataset wrappers, dump first table for diagnosis
    if all_tables and not all_market:
        print("Sample of first table:")
        rows = all_tables[0].find_all("tr")[:5]
        for r in rows:
            print("  ", [td.get_text(strip=True) for td in r.find_all(["td", "th"])])

    results = {}

    for section in all_market:
        h2 = section.find("h2")
        if not h2:
            continue
        location_name = h2.get_text(strip=True)
        if location_name not in LOCATIONS:
            continue

        location_data = {}

        for product_section in section.find_all("div", class_="dataset"):
            parent = product_section.parent
            h3 = parent.find("h3") if parent else None
            if not h3:
                continue
            raw_name = h3.get_text(separator=" ", strip=True)
            product_name = raw_name.replace("ED", "").split("\n")[0].strip()

            if not any(pt.lower() in product_name.lower() for pt in PRODUCT_TYPES):
                continue

            table = product_section.find("table")
            if not table:
                continue

            thead = table.find("thead")
            if thead:
                header_cells = [th.get_text(strip=True) for th in thead.find_all("th")]
            else:
                header_cells = ["Contract"] + CONTRACT_FIELDS

            rows = []
            tbody = table.find("tbody")
            if tbody:
                for tr in tbody.find_all("tr"):
                    cells = [td.get_text(strip=True) for td in tr.find_all("td")]
                    if cells and any(c for c in cells):
                        row_dict = {}
                        for i, val in enumerate(cells):
                            col_name = header_cells[i] if i < len(header_cells) else f"Col{i}"
                            row_dict[col_name] = val
                        rows.append(row_dict)

            if rows:
                location_data[product_name] = rows
                print(f"  Parsed {location_name} / {product_name}: {len(rows)} contracts")

        results[location_name] = location_data

    return last_update, results


def get_nzt_timestamp():
    nzt = pytz.timezone("Pacific/Auckland")
    return datetime.now(nzt).strftime("%Y-%m-%d %H:%M:%S NZT")


def build_all_headers(data):
    """
    Build the full ordered column list for the flat one-row-per-day format.
    Columns: Scraped At | ASX Last Update | Location | Product | Contract | Field ...
    """
    cols = ["Scraped At", "ASX Last Update"]
    for location in LOCATIONS:
        if location not in data:
            continue
        for product_name, rows in data[location].items():
            for row_dict in rows:
                contract = row_dict.get("Contract", "")
                for field in CONTRACT_FIELDS:
                    col = f"{location} | {product_name} | {contract} | {field}"
                    if col not in cols:
                        cols.append(col)
    return cols


def write_to_excel(last_update, data):
    scraped_at = get_nzt_timestamp()

    # Build flat dict for this scrape: column_name -> value
    flat_row = {"Scraped At": scraped_at, "ASX Last Update": last_update}

    for location in LOCATIONS:
        if location not in data:
            continue
        for product_name, rows in data[location].items():
            for row_dict in rows:
                contract = row_dict.get("Contract", "")
                for field in CONTRACT_FIELDS:
                    col = f"{location} | {product_name} | {contract} | {field}"
                    flat_row[col] = row_dict.get(field, "")

    sheet_name = "Data"

    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            # Remove default blank sheet if present
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

    # Read existing headers from row 1
    existing_headers = [cell.value for cell in ws[1] if cell.value is not None]

    # Merge with any new columns from today's data
    new_cols = build_all_headers(data)
    all_headers = list(existing_headers)
    for col in new_cols:
        if col not in all_headers:
            all_headers.append(col)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    if not existing_headers:
        # New sheet — write full header row
        for col_idx, col_name in enumerate(all_headers, 1):
            cell = ws.cell(1, col_idx, col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[1].height = 60
        ws.freeze_panes = "C2"
    elif len(all_headers) > len(existing_headers):
        # New contracts appeared — extend header row
        for col_idx in range(len(existing_headers) + 1, len(all_headers) + 1):
            cell = ws.cell(1, col_idx, all_headers[col_idx - 1])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Set column widths
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 34
    for col_idx in range(3, len(all_headers) + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 14

    # Append the data row
    next_row = ws.max_row + 1
    for col_idx, col_name in enumerate(all_headers, 1):
        ws.cell(next_row, col_idx, flat_row.get(col_name, ""))

    wb.save(EXCEL_FILE)
    print(f"Saved to {EXCEL_FILE} — {scraped_at} | ASX Update: {last_update}")
    print(f"Total columns: {len(all_headers)} | Row written: {next_row}")


def main():
    print("Fetching ASX Energy NZ futures data (headless Chrome)...")
    html = fetch_html_with_selenium()
    last_update, data = parse_data(html)
    print(f"ASX Last Update: {last_update}")
    write_to_excel(last_update, data)


if __name__ == "__main__":
    main()
